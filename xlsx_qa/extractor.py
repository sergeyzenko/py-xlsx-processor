"""Question extraction logic."""

from __future__ import annotations

from typing import Dict, List, Optional

from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from .domain import Question

QUESTION_WORDS = (
    "who",
    "what",
    "when",
    "where",
    "why",
    "how",
    "do",
    "does",
    "did",
    "is",
    "are",
    "can",
    "could",
    "will",
    "would",
    "please",
)

PATTERN_HINTS = (
    "please describe",
    "please explain",
    "provide details",
)


class QuestionExtractor:
    """Scan a workbook and build question objects."""

    def extract(self, workbook: Workbook) -> List[Question]:
        questions: List[Question] = []
        question_id = 1

        for sheet in workbook.worksheets:
            merged_lookup = _build_merged_lookup(sheet)

            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue

                    if cell.coordinate in merged_lookup.skip_coordinates:
                        continue

                    text = _normalise(cell.value)
                    if not text:
                        continue

                    if not self._is_question_text(text):
                        continue

                    answer_cell = _resolve_answer_cell(sheet, cell, merged_lookup)

                    preexisting = _safe_value(answer_cell) if answer_cell else None

                    questions.append(
                        Question(
                            id=question_id,
                            sheet=sheet.title,
                            coord=cell.coordinate,
                            row=cell.row,
                            col=cell.column,
                            text=text,
                            answer_coord=answer_cell.coordinate if answer_cell else get_followup_coordinate(cell),
                            preexisting_answer=preexisting,
                        )
                    )
                    question_id += 1

        return questions

    def _is_question_text(self, text: str) -> bool:
        """Heuristic classifier that determines if text is a question prompt."""
        lowered = text.lower()
        if lowered.endswith("?"):
            return True

        if len(lowered) < 10:
            return False

        if any(lowered.startswith(word) for word in QUESTION_WORDS):
            return True

        return any(hint in lowered for hint in PATTERN_HINTS)


def _normalise(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return ""
    return str(value).strip()


def _safe_value(cell: Optional[Cell]) -> Optional[str]:
    if not cell:
        return None
    value = cell.value
    if value is None:
        return None
    return str(value)


class _MergedLookup:
    """Helper structure for merged-cell calculations."""

    __slots__ = ("rightmost_column", "skip_coordinates")

    def __init__(self, rightmost_column: Dict[str, int], skip_coordinates: set[str]) -> None:
        self.rightmost_column = rightmost_column
        self.skip_coordinates = skip_coordinates


def _build_merged_lookup(sheet: Worksheet) -> _MergedLookup:
    rightmost_column: Dict[str, int] = {}
    skip_coordinates: set[str] = set()

    for merged_range in sheet.merged_cells.ranges:
        min_row, max_row = merged_range.min_row, merged_range.max_row
        min_col, max_col = merged_range.min_col, merged_range.max_col
        top_left = sheet.cell(row=min_row, column=min_col)
        rightmost_column[top_left.coordinate] = max_col

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                if coord != top_left.coordinate:
                    skip_coordinates.add(coord)

    return _MergedLookup(rightmost_column, skip_coordinates)


def _resolve_answer_cell(sheet: Worksheet, question_cell: Cell, merged_lookup: _MergedLookup) -> Optional[Cell]:
    rightmost_col = merged_lookup.rightmost_column.get(question_cell.coordinate, question_cell.column)
    target_column = rightmost_col + 1
    return sheet.cell(row=question_cell.row, column=target_column)


def get_followup_coordinate(question_cell: Cell) -> str:
    rightmost_col = question_cell.column
    target_column = rightmost_col + 1
    return f"{get_column_letter(target_column)}{question_cell.row}"
